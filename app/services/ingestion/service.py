"""Admin ingestion service for topic, question, rubric, and context imports."""

from __future__ import annotations

import csv
import io
from pathlib import Path
import textwrap

from openpyxl import load_workbook
from pypdf import PdfReader

from app.models.schemas.admin import (
    ImportApplyResponse,
    ImportMessage,
    ImportPreviewResponse,
    TopicPackageImportModel,
    TopicSummaryResponse,
)
from app.providers.file_storage import LocalFileStorage
from app.repositories.content_repository import SqliteContentRepository
from app.services.retrieval.service import RetrievalService


class IngestionError(ValueError):
    """Raised when import validation or persistence fails."""


class IngestionService:
    """Validates and persists admin-managed content imports."""

    def __init__(
        self,
        content_repository: SqliteContentRepository,
        retrieval_service: RetrievalService,
        file_storage: LocalFileStorage,
    ) -> None:
        self._content_repository = content_repository
        self._retrieval_service = retrieval_service
        self._file_storage = file_storage

    def list_topics(self, include_unpublished: bool = False) -> list[TopicSummaryResponse]:
        topics = self._content_repository.list_topics(include_unpublished=include_unpublished)
        return [
            TopicSummaryResponse(
                id=topic.id,
                topic_code=topic.topic_code,
                topic_name=topic.topic_name,
                description=topic.description,
                published=topic.published,
                question_count=topic.question_count,
            )
            for topic in topics
        ]

    def set_topic_publication(self, topic_id: str, published: bool):
        topic = self._content_repository.set_topic_publication(topic_id, published)
        if topic is None:
            raise IngestionError("Topic not found")
        return topic

    def list_questions(self, topic_id: str, published_only: bool = False):
        return self._content_repository.list_questions_by_topic(topic_id, published_only=published_only)

    def get_rubric(self, question_id: str):
        return self._content_repository.list_rubric_criteria(question_id)

    def get_weight_config(self, question_id: str):
        return self._content_repository.get_weight_config(question_id)

    def update_rubric_weights(self, weights_by_criterion_id: dict[str, float]) -> None:
        for criterion_id, weight in weights_by_criterion_id.items():
            self._content_repository.update_rubric_criterion_weight(criterion_id, weight)

    def update_weight_config(
        self,
        question_id: str,
        *,
        confidence_low: float,
        confidence_mid_start: float,
        confidence_mid_end: float,
        confidence_high: float,
        max_followups: int,
    ):
        return self._content_repository.upsert_weight_config(
            question_id,
            confidence_low=confidence_low,
            confidence_mid_start=confidence_mid_start,
            confidence_mid_end=confidence_mid_end,
            confidence_high=confidence_high,
            max_followups=max_followups,
        )

    def preview_uploaded_file(
        self,
        file_name: str,
        file_bytes: bytes,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> ImportPreviewResponse:
        """Preview a structured or PDF upload from raw bytes."""

        suffix = Path(file_name).suffix.lower()
        if suffix == ".json":
            package = TopicPackageImportModel.model_validate_json(file_bytes.decode("utf-8"))
            return self.preview_json_package(package)
        if suffix == ".csv":
            return self.preview_csv_text(file_bytes.decode("utf-8"), topic_code=topic_code, topic_name=topic_name)
        if suffix in {".xlsx", ".xlsm"}:
            return self.preview_excel_bytes(file_bytes, topic_code=topic_code, topic_name=topic_name)
        if suffix == ".pdf":
            return self.preview_pdf_bytes(file_bytes)
        raise IngestionError("Unsupported upload type")

    def apply_uploaded_file(
        self,
        *,
        file_name: str,
        file_bytes: bytes,
        created_by: str,
        topic_code: str | None = None,
        topic_name: str | None = None,
        question_code: str | None = None,
        published: bool = False,
    ) -> ImportApplyResponse:
        """Persist a structured or PDF upload from raw bytes."""

        suffix = Path(file_name).suffix.lower()
        storage_ref = self._file_storage.save_upload(file_name, file_bytes)
        if suffix == ".json":
            package = TopicPackageImportModel.model_validate_json(file_bytes.decode("utf-8"))
            return self.apply_json_package(package, created_by, source_reference=storage_ref)
        if suffix == ".csv":
            return self.apply_csv_text(file_bytes.decode("utf-8"), created_by, source_reference=storage_ref, topic_code=topic_code, topic_name=topic_name)
        if suffix in {".xlsx", ".xlsm"}:
            return self.apply_excel_bytes(file_bytes, created_by, source_reference=storage_ref, topic_code=topic_code, topic_name=topic_name)
        if suffix == ".pdf":
            if not topic_code:
                raise IngestionError("PDF imports require a topic code")
            return self.apply_pdf_bytes(
                file_name=file_name,
                file_bytes=file_bytes,
                created_by=created_by,
                topic_code=topic_code,
                question_code=question_code,
                published=published,
                storage_ref=storage_ref,
            )
        raise IngestionError("Unsupported upload type")

    def preview_json_package(self, package: TopicPackageImportModel) -> ImportPreviewResponse:
        messages = self._validate_json_package(package)
        return ImportPreviewResponse(
            valid=not any(message.severity == "error" for message in messages),
            messages=messages,
            topic_count=1,
            question_count=len(package.questions),
            context_count=len(package.contexts),
            rubric_count=sum(len(question.rubric) for question in package.questions),
        )

    def apply_json_package(
        self,
        package: TopicPackageImportModel,
        created_by: str,
        source_reference: str | None = None,
    ) -> ImportApplyResponse:
        preview = self.preview_json_package(package)
        if not preview.valid:
            raise IngestionError("JSON package contains validation errors")

        existing_topic = self._content_repository.get_topic_by_code(package.topic.topic_code)
        if existing_topic is not None:
            raise IngestionError("Topic code already exists")

        topic = self._content_repository.create_topic(
            topic_code=package.topic.topic_code,
            topic_name=package.topic.topic_name,
            description=package.topic.description,
            created_by=created_by,
            published=package.topic.published,
        )

        question_lookup: dict[str, str] = {}
        rubric_count = 0
        for question_payload in package.questions:
            question = self._content_repository.create_question(
                topic_id=topic.id,
                question_code=question_payload.question_code,
                question_text=question_payload.question_text,
                question_type=question_payload.question_type,
                difficulty=question_payload.difficulty,
                expected_answer_summary=question_payload.expected_answer_summary,
                followup_enabled=question_payload.followup_enabled,
                published=topic.published if question_payload.published is None else question_payload.published,
                prompt_notes=question_payload.question_prompt_notes,
                time_limit_seconds=question_payload.time_limit_seconds,
                tags=question_payload.tags,
                language=question_payload.language,
                source_reference=question_payload.source_reference or source_reference,
            )
            question_lookup[question_payload.question_code] = question.id
            self._content_repository.upsert_weight_config(
                question_id=question.id,
                confidence_low=question_payload.confidence_thresholds["low"],
                confidence_mid_start=question_payload.confidence_thresholds["mid_start"],
                confidence_mid_end=question_payload.confidence_thresholds["mid_end"],
                confidence_high=question_payload.confidence_thresholds["high"],
                max_followups=question_payload.max_followups,
            )
            for criterion in question_payload.rubric:
                self._content_repository.create_rubric_criterion(
                    question_id=question.id,
                    criterion_code=criterion.criterion_code,
                    criterion_name=criterion.criterion_name,
                    criterion_description=criterion.criterion_description,
                    weight=criterion.weight,
                    min_score=criterion.min_score,
                    max_score=criterion.max_score,
                    evidence_required=criterion.evidence_required,
                )
                rubric_count += 1

        context_count = 0
        created_contexts = []
        for context_payload in package.contexts:
            question_id = question_lookup.get(context_payload.scope_code) if context_payload.scope_type == "question" else None
            created_context = self._content_repository.create_context(
                topic_id=topic.id,
                question_id=question_id,
                context_code=context_payload.context_code,
                source_type=context_payload.source_type,
                context_title=context_payload.context_title,
                context_text=context_payload.context_text,
                storage_ref=source_reference,
                page_reference=context_payload.page_reference,
                section_reference=context_payload.section_reference,
                priority=context_payload.priority,
                published=topic.published if context_payload.published is None else context_payload.published,
            )
            created_contexts.append(created_context)
            context_count += 1

        self._retrieval_service.index_contexts(created_contexts)

        return ImportApplyResponse(
            topic_id=topic.id,
            topic_code=topic.topic_code,
            topic_name=topic.topic_name,
            question_count=len(package.questions),
            context_count=context_count,
            rubric_count=rubric_count,
            published=topic.published,
        )

    def preview_csv_text(
        self,
        csv_text: str,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> ImportPreviewResponse:
        rows, messages = self._parse_csv_rows(csv_text, topic_code=topic_code, topic_name=topic_name)
        return ImportPreviewResponse(
            valid=not any(message.severity == "error" for message in messages),
            messages=messages,
            topic_count=len({row.get("topic_code", "") for row in rows} - {""}) if rows else (1 if topic_code else 0),
            question_count=len(rows),
            context_count=_estimate_question_context_count(rows),
            rubric_count=len(rows),
        )

    def preview_excel_bytes(
        self,
        file_bytes: bytes,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> ImportPreviewResponse:
        rows, messages = self._parse_excel_rows(file_bytes, topic_code=topic_code, topic_name=topic_name)
        return ImportPreviewResponse(
            valid=not any(message.severity == "error" for message in messages),
            messages=messages,
            topic_count=len({row.get("topic_code", "") for row in rows} - {""}) if rows else (1 if topic_code else 0),
            question_count=len(rows),
            context_count=_estimate_question_context_count(rows),
            rubric_count=len(rows),
        )

    def apply_csv_text(
        self,
        csv_text: str,
        created_by: str,
        source_reference: str | None = None,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> ImportApplyResponse:
        rows, messages = self._parse_csv_rows(csv_text, topic_code=topic_code, topic_name=topic_name)
        return self._apply_tabular_rows(rows, messages, created_by, source_reference, topic_code=topic_code, topic_name=topic_name)

    def apply_excel_bytes(
        self,
        file_bytes: bytes,
        created_by: str,
        source_reference: str | None = None,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> ImportApplyResponse:
        rows, messages = self._parse_excel_rows(file_bytes, topic_code=topic_code, topic_name=topic_name)
        return self._apply_tabular_rows(rows, messages, created_by, source_reference, topic_code=topic_code, topic_name=topic_name)

    def preview_pdf_bytes(self, file_bytes: bytes) -> ImportPreviewResponse:
        chunks = self._extract_pdf_chunks(file_bytes)
        messages = [] if chunks else [ImportMessage(severity="error", message="PDF does not contain extractable text")]
        return ImportPreviewResponse(
            valid=bool(chunks),
            messages=messages,
            topic_count=0,
            question_count=0,
            context_count=len(chunks),
            rubric_count=0,
        )

    def apply_pdf_bytes(
        self,
        *,
        file_name: str,
        file_bytes: bytes,
        created_by: str,
        topic_code: str,
        question_code: str | None,
        published: bool,
        storage_ref: str | None,
    ) -> ImportApplyResponse:
        del created_by
        topic = self._content_repository.get_topic_by_code(topic_code)
        if topic is None:
            raise IngestionError("Topic code not found for PDF import")

        question_id = None
        if question_code:
            question = self._content_repository.get_question_by_code(question_code)
            if question is None:
                raise IngestionError("Question code not found for PDF import")
            question_id = question.id

        chunks = self._extract_pdf_chunks(file_bytes)
        if not chunks:
            raise IngestionError("PDF text content is empty")

        created_contexts = []
        for index, chunk in enumerate(chunks, start=1):
            created_contexts.append(
                self._content_repository.create_context(
                    topic_id=topic.id,
                    question_id=question_id,
                    context_code=f"{topic_code}_PDF_{index}",
                    source_type="pdf",
                    context_title=f"{file_name} part {index}",
                    context_text=chunk["text"],
                    storage_ref=storage_ref,
                    page_reference=chunk["page_reference"],
                    section_reference=None,
                    priority=1,
                    published=published,
                )
            )

        self._retrieval_service.index_contexts(created_contexts)
        return ImportApplyResponse(
            topic_id=topic.id,
            topic_code=topic.topic_code,
            topic_name=topic.topic_name,
            question_count=topic.question_count,
            context_count=len(created_contexts),
            rubric_count=0,
            published=topic.published,
        )

    def _apply_tabular_rows(
        self,
        rows: list[dict[str, str]],
        messages: list[ImportMessage],
        created_by: str,
        source_reference: str | None,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> ImportApplyResponse:
        if any(message.severity == "error" for message in messages):
            raise IngestionError("Tabular import contains validation errors")
        if not rows:
            raise IngestionError("Tabular import is empty")

        # UI overrides take precedence over CSV column values
        topic_code = topic_code or rows[0].get("topic_code", "")
        topic_name = topic_name or rows[0].get("topic_name", "")
        if self._content_repository.get_topic_by_code(topic_code) is not None:
            raise IngestionError("Topic code already exists")

        topic = self._content_repository.create_topic(
            topic_code=topic_code,
            topic_name=topic_name,
            description=f"Imported question bank for {topic_name}",
            created_by=created_by,
            published=_parse_bool(rows[0].get("published", "false")),
        )

        rubric_count = 0
        context_count = 0
        created_contexts = []
        for row in rows:
            question = self._content_repository.create_question(
                topic_id=topic.id,
                question_code=row["question_code"],
                question_text=row["question_text"],
                question_type=row["question_type"],
                difficulty=row["difficulty"],
                expected_answer_summary=row["expected_answer_summary"],
                followup_enabled=_parse_bool(row["followup_enabled"]),
                published=_parse_bool(row["published"]),
                prompt_notes=row.get("question_prompt_notes") or None,
                time_limit_seconds=int(row["time_limit_seconds"]) if row.get("time_limit_seconds") else None,
                tags=_split_pipe_values(row.get("tags", "")),
                language=row.get("language") or None,
                source_reference=row.get("source_reference") or source_reference,
            )
            question_chunks = _build_question_context_chunks(question.question_text)
            for chunk_index, chunk_text in enumerate(question_chunks, start=1):
                title_suffix = f" part {chunk_index}" if len(question_chunks) > 1 else ""
                created_context = self._content_repository.create_context(
                    topic_id=topic.id,
                    question_id=question.id,
                    context_code=f"{question.question_code}_QUESTION_{chunk_index}",
                    source_type="question",
                    context_title=f"Question text for {question.question_code}{title_suffix}",
                    context_text=chunk_text,
                    storage_ref=source_reference,
                    page_reference=None,
                    section_reference=None,
                    priority=1,
                    published=question.published,
                )
                created_contexts.append(created_context)
                context_count += 1
            self._content_repository.create_rubric_criterion(
                question_id=question.id,
                criterion_code=f"{question.question_code}_COVERAGE",
                criterion_name="Expected Coverage",
                criterion_description=question.expected_answer_summary,
                weight=1.0,
                min_score=0,
                max_score=5,
                evidence_required=True,
            )
            rubric_count += 1
            self._content_repository.upsert_weight_config(
                question_id=question.id,
                confidence_low=float(row["confidence_low"]),
                confidence_mid_start=float(row["confidence_mid_start"]),
                confidence_mid_end=float(row["confidence_mid_end"]),
                confidence_high=float(row["confidence_high"]),
                max_followups=int(row["max_followups"]),
            )

        self._retrieval_service.index_contexts(created_contexts)

        return ImportApplyResponse(
            topic_id=topic.id,
            topic_code=topic.topic_code,
            topic_name=topic.topic_name,
            question_count=len(rows),
            context_count=context_count,
            rubric_count=rubric_count,
            published=topic.published,
        )

    def _validate_json_package(self, package: TopicPackageImportModel) -> list[ImportMessage]:
        messages: list[ImportMessage] = []
        question_codes = set()
        context_codes = set()
        for index, question in enumerate(package.questions, start=1):
            if question.question_code in question_codes:
                messages.append(ImportMessage(severity="error", field_name="question_code", row_number=index, message="Duplicate question code"))
            if self._content_repository.get_question_by_code(question.question_code) is not None:
                messages.append(
                    ImportMessage(
                        severity="error",
                        field_name="question_code",
                        row_number=index,
                        message=f"Question code '{question.question_code}' already exists in the system",
                    )
                )
            question_codes.add(question.question_code)
            thresholds = question.confidence_thresholds
            if not (
                thresholds["low"] <= thresholds["mid_start"] <= thresholds["mid_end"] <= thresholds["high"]
            ):
                messages.append(ImportMessage(severity="error", field_name="confidence_thresholds", row_number=index, message="Confidence thresholds are invalid"))
            total_weight = sum(criterion.weight for criterion in question.rubric)
            if round(total_weight, 5) not in {1.0, 100.0}:
                messages.append(ImportMessage(severity="error", field_name="rubric.weight", row_number=index, message="Rubric weights must sum to 1.0 or 100"))
            for context_code in question.context_codes:
                if context_code not in {context.context_code for context in package.contexts}:
                    messages.append(ImportMessage(severity="error", field_name="context_codes", row_number=index, message=f"Unknown context code '{context_code}' referenced by question"))

        for index, context in enumerate(package.contexts, start=1):
            if context.context_code in context_codes:
                messages.append(ImportMessage(severity="error", field_name="context_code", row_number=index, message="Duplicate context code"))
            context_codes.add(context.context_code)
            if context.scope_type == "question" and context.scope_code not in question_codes:
                messages.append(ImportMessage(severity="error", field_name="scope_code", row_number=index, message="Context references an unknown question code"))

        return messages

    def _parse_csv_rows(
        self,
        csv_text: str,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> tuple[list[dict[str, str]], list[ImportMessage]]:
        reader = csv.DictReader(io.StringIO(csv_text))
        if reader.fieldnames is None:
            return [], [ImportMessage(severity="error", message="CSV content is missing headers")]

        rows = [{key: (value or "").strip() for key, value in row.items()} for row in reader]
        return rows, self._validate_tabular_rows(rows, set(reader.fieldnames), topic_code=topic_code, topic_name=topic_name)

    def _parse_excel_rows(
        self,
        file_bytes: bytes,
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> tuple[list[dict[str, str]], list[ImportMessage]]:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        if "questions" not in workbook.sheetnames:
            return [], [ImportMessage(severity="error", field_name="worksheet", message="Excel import must contain a 'questions' worksheet")]

        worksheet = workbook["questions"]
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(row_iter)]
        except StopIteration:
            return [], [ImportMessage(severity="error", message="Excel worksheet is empty")]

        rows: list[dict[str, str]] = []
        for row in row_iter:
            rows.append(
                {
                    headers[index]: "" if value is None else str(value).strip()
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
            )

        return rows, self._validate_tabular_rows(rows, set(headers), topic_code=topic_code, topic_name=topic_name)

    def _validate_tabular_rows(
        self,
        rows: list[dict[str, str]],
        headers: set[str],
        topic_code: str | None = None,
        topic_name: str | None = None,
    ) -> list[ImportMessage]:
        messages: list[ImportMessage] = []
        # topic_code / topic_name are not required in the file when supplied as UI overrides
        override_cols: set[str] = set()
        if topic_code:
            override_cols.add("topic_code")
        if topic_name:
            override_cols.add("topic_name")
        required_columns = {
            "topic_code",
            "topic_name",
            "question_code",
            "question_text",
            "question_type",
            "difficulty",
            "expected_answer_summary",
            "followup_enabled",
            "max_followups",
            "confidence_low",
            "confidence_mid_start",
            "confidence_mid_end",
            "confidence_high",
            "published",
        } - override_cols
        missing_columns = required_columns.difference(headers)
        for missing in sorted(missing_columns):
            messages.append(ImportMessage(severity="error", field_name=missing, message="Required column is missing"))

        # only validate per-row values for columns that are actually present
        present_required = required_columns.intersection(headers)
        seen_question_codes: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            for column in present_required:
                if not row.get(column):
                    messages.append(ImportMessage(severity="error", field_name=column, row_number=row_number, message="Required value is empty"))
            if row.get("question_code") in seen_question_codes:
                messages.append(ImportMessage(severity="error", field_name="question_code", row_number=row_number, message="Duplicate question code"))
            question_code = row.get("question_code", "")
            seen_question_codes.add(question_code)
            if question_code and self._content_repository.get_question_by_code(question_code) is not None:
                messages.append(
                    ImportMessage(
                        severity="error",
                        field_name="question_code",
                        row_number=row_number,
                        message=f"Question code '{question_code}' already exists in the system",
                    )
                )
            # skip per-row threshold check when columns are already flagged missing
            if all(row.get(c) for c in ("confidence_low", "confidence_mid_start", "confidence_mid_end", "confidence_high")):
                try:
                    low = float(row["confidence_low"])
                    mid_start = float(row["confidence_mid_start"])
                    mid_end = float(row["confidence_mid_end"])
                    high = float(row["confidence_high"])
                    if not (low <= mid_start <= mid_end <= high):
                        raise ValueError()
                except ValueError:
                    messages.append(ImportMessage(severity="error", field_name="confidence_low", row_number=row_number, message="Confidence thresholds are invalid"))
            if row.get("max_followups"):
                try:
                    max_followups = int(row["max_followups"])
                    if max_followups < 0 or max_followups > 3:
                        raise ValueError()
                except ValueError:
                    messages.append(ImportMessage(severity="error", field_name="max_followups", row_number=row_number, message="Max followups must be between 0 and 3"))
            for field_name in ("followup_enabled", "published"):
                if row.get(field_name):
                    try:
                        _parse_bool(row[field_name])
                    except IngestionError:
                        messages.append(ImportMessage(severity="error", field_name=field_name, row_number=row_number, message="Boolean values must be true or false"))

        return messages

    def _extract_pdf_chunks(self, file_bytes: bytes) -> list[dict[str, str]]:
        reader = PdfReader(io.BytesIO(file_bytes))
        chunks: list[dict[str, str]] = []
        for page_number, page in enumerate(reader.pages, start=1):
            page_text = (page.extract_text() or "").strip()
            if not page_text:
                continue
            for chunk_index, chunk in enumerate(_chunk_text(page_text), start=1):
                chunks.append(
                    {
                        "text": chunk,
                        "page_reference": f"page-{page_number}-chunk-{chunk_index}",
                    }
                )
        return chunks


def _parse_bool(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized not in {"true", "false"}:
        raise IngestionError("Boolean values must be true or false")
    return normalized == "true"


def _split_pipe_values(value: str) -> list[str]:
    return [item.strip() for item in value.split("|") if item.strip()]


def _build_question_context_chunks(question_text: str) -> list[str]:
    """Build retrieval chunks from question text only, excluding expected answers."""

    return _chunk_text(question_text, width=320, overlap=48)


def _estimate_question_context_count(rows: list[dict[str, str]]) -> int:
    """Estimate how many question-derived chunks will be created during tabular import."""

    return sum(len(_build_question_context_chunks(row.get("question_text", ""))) for row in rows)


def _chunk_text(text: str, width: int = 900, overlap: int = 150) -> list[str]:
    normalized = " ".join(text.split())
    if not normalized:
        return []
    if len(normalized) <= width:
        return [normalized]

    chunks: list[str] = []
    start = 0
    while start < len(normalized):
        end = min(start + width, len(normalized))
        chunks.append(normalized[start:end])
        if end >= len(normalized):
            break
        start = max(end - overlap, start + 1)
    return chunks
